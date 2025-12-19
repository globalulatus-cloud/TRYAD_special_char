import streamlit as st
import openpyxl
import os
import tempfile

# --------------------------------------------------
# Page configuration
# --------------------------------------------------
st.set_page_config(
    page_title="Excel JP Punctuation Validator",
    layout="centered"
)

st.title("Japanese Punctuation Validator and Annotator")
st.write(
    "Upload an Excel file to validate Japanese punctuation against the source "
    "and download an annotated version."
)

# --------------------------------------------------
# Character maps
# --------------------------------------------------
HALF_TO_FULL_MAP = {
    "(": "（",
    ")": "）",
    "[": "［",
    "]": "］",
    ",": "、",
    "/": "／",
    ".": "。",
    "X": "×",
    ":": "：",
    "#": "＃",
}

FULL_TO_HALF_MAP = {v: k for k, v in HALF_TO_FULL_MAP.items()}

# --------------------------------------------------
# Core processing function
# --------------------------------------------------
def process_and_validate_excel(file_path, h2f_map, f2h_map):
    workbook = openpyxl.load_workbook(file_path)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        sheet.cell(row=1, column=3, value="Fixed Japanese")
        sheet.cell(row=1, column=4, value="Reason for Change / Validation")

        for row in sheet.iter_rows(min_row=2):
            source_cell = row[0]  # Column A
            target_cell = row[1]  # Column B

            source_text = source_cell.value if isinstance(source_cell.value, str) else ""
            original_target_text = target_cell.value if isinstance(target_cell.value, str) else ""

            reasons = []
            fixed_target_text = original_target_text

            # Fix half-width to full-width
            for half, full in h2f_map.items():
                if half in fixed_target_text:
                    fixed_target_text = fixed_target_text.replace(half, full)
                    reasons.append(f"Replaced '{half}' with '{full}'")

            # Missing punctuation
            for half, full in h2f_map.items():
                if half in source_text and full not in original_target_text:
                    msg = f"Missing: '{full}'"
                    if msg not in reasons:
                        reasons.append(msg)

            # Additional punctuation
            for full, half in f2h_map.items():
                if full in original_target_text and half not in source_text:
                    reasons.append(f"Additional: '{full}'")

            sheet.cell(row=target_cell.row, column=3, value=fixed_target_text)

            if reasons:
                sheet.cell(
                    row=target_cell.row,
                    column=4,
                    value="; ".join(sorted(set(reasons)))
                )

    output_path = file_path.replace(".xlsx", "_validated.xlsx")
    workbook.save(output_path)
    return output_path

# --------------------------------------------------
# Upload UI
# --------------------------------------------------
uploaded_file = st.file_uploader(
    "Upload Excel file (.xlsx)",
    type=["xlsx"]
)

if uploaded_file:
    st.success("File uploaded successfully")

    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = os.path.join(tmpdir, uploaded_file.name)

        with open(input_path, "wb") as f:
            f.write(uploaded_file.read())

        if st.button("Process File"):
            with st.spinner("Processing and validating file..."):
                try:
                    output_path = process_and_validate_excel(
                        input_path,
                        HALF_TO_FULL_MAP,
                        FULL_TO_HALF_MAP
                    )

                    with open(output_path, "rb") as f:
                        output_bytes = f.read()

                    st.success("Processing complete")

                    st.download_button(
                        label="Download validated Excel file",
                        data=output_bytes,
                        file_name=os.path.basename(output_path),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                except Exception as e:
                    st.error(f"Processing failed: {e}")
